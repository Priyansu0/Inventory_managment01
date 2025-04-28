"""
Utilities for exporting data to Excel and CSV.
"""

import os
import logging
import csv
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

logger = logging.getLogger(__name__)


def export_to_excel(file_path, data, sheet_name=None, headers=None, rows=None):
    """Export data to Excel file.
    
    Args:
        file_path (str): Path where to save the Excel file
        data: Either a dict with sheet data or the data rows
        sheet_name (str, optional): Name of the sheet (only used if data is not a dict)
        headers (list, optional): List of column headers (only used if data is not a dict)
        rows (list, optional): List of data rows (only used if data is not a dict)
    """
    try:
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        
        # Check if data is a dict with multiple sheets
        if isinstance(data, dict):
            if "chart_path" in data:
                chart_path = data.pop("chart_path")
            else:
                chart_path = None
                
            # Process each sheet in the dict
            for sheet_name, sheet_data in data.items():
                if isinstance(sheet_data, dict):
                    # Format: {"headers": [...], "data": [...]}
                    headers = sheet_data.get("headers", [])
                    rows = sheet_data.get("data", [])
                else:
                    # Assume sheet_data is the row data
                    rows = sheet_data
                    headers = []
                
                create_excel_sheet(workbook, sheet_name, headers, rows)
            
            # Add chart if provided
            if chart_path and os.path.exists(chart_path):
                try:
                    # Add a chart sheet
                    chart_sheet = workbook.create_sheet("Chart")
                    img = XLImage(chart_path)
                    
                    # Calculate cell position (centered)
                    img_width_px = img.width
                    img_height_px = img.height
                    
                    # Scale down if too large
                    max_width = 800
                    max_height = 600
                    
                    if img_width_px > max_width or img_height_px > max_height:
                        scale = min(max_width / img_width_px, max_height / img_height_px)
                        img.width = int(img_width_px * scale)
                        img.height = int(img_height_px * scale)
                    
                    # Add image to sheet
                    chart_sheet.add_image(img, 'B2')
                    
                    # Delete temp file
                    os.unlink(chart_path)
                except Exception as e:
                    logger.error(f"Error adding chart to Excel: {str(e)}")
            
        else:
            # Single sheet format
            create_excel_sheet(workbook, sheet_name or "Sheet1", headers or [], data)
        
        # Save the workbook
        workbook.save(file_path)
        logger.info(f"Data exported to Excel file: {file_path}")
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        raise


def create_excel_sheet(workbook, sheet_name, headers, data):
    """Create and populate a sheet in the Excel workbook.
    
    Args:
        workbook: The openpyxl Workbook
        sheet_name (str): Name of the sheet
        headers (list): List of column headers
        data (list): List of data rows
    """
    # Create sheet
    sheet = workbook.create_sheet(title=sheet_name)
    
    # Add headers
    if headers:
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add data rows
    for row_idx, row_data in enumerate(data, 2 if headers else 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Auto-adjust column widths
    for col_idx, _ in enumerate(headers if headers else (data[0] if data else []), 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        
        for row_idx in range(1, len(data) + (2 if headers else 1)):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        adjusted_width = max(max_length + 2, 10)  # Min width of 10
        sheet.column_dimensions[col_letter].width = min(adjusted_width, 50)  # Max width of 50
    
    # Freeze header row if we have headers
    if headers:
        sheet.freeze_panes = "A2"


def export_to_csv(file_path, headers, data):
    """Export data to CSV file.
    
    Args:
        file_path (str): Path where to save the CSV file
        headers (list): List of column headers
        data (list): List of data rows
    """
    try:
        with open(file_path, 'w', newline='', encoding='utf-8') as csv_file:
            writer = csv.writer(csv_file)
            
            # Write headers
            if headers:
                writer.writerow(headers)
            
            # Write data
            for row in data:
                writer.writerow(row)
        
        logger.info(f"Data exported to CSV file: {file_path}")
        
    except Exception as e:
        logger.error(f"Error exporting to CSV: {str(e)}")
        raise


def format_datetime(dt):
    """Format a datetime object for export.
    
    Args:
        dt: Datetime object or None
        
    Returns:
        str: Formatted date string or empty string if None
    """
    if not dt:
        return ""
    
    if isinstance(dt, datetime.datetime):
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    elif isinstance(dt, datetime.date):
        return dt.strftime("%Y-%m-%d")
    else:
        return str(dt)
